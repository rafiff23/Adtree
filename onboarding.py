import re
import streamlit as st
import pandas as pd
import io
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from db import get_connection

# ── XLSX template helpers ─────────────────────────────────────────────────────
REGISTRY_COLUMNS = [
    "tiktok_id", "followers", "full_name", "domicile",
    "uid", "phone_number", "tiktok_link", "binding_status",
    "onboarding_date", "level",
]

# Columns for the unmatched download — stripped to only what needs manual input
UNMATCHED_COLUMNS = ["tiktok_id", "full_name", "domicile", "uid", "phone_number"]

HEADER_FILL = PatternFill("solid", start_color="1E293B", end_color="1E293B")
HEADER_FONT = Font(bold=True, color="FFFFFF", name="Arial", size=10)
CELL_FONT   = Font(name="Arial", size=10)
BORDER_SIDE = Side(style="thin", color="CBD5E1")
CELL_BORDER = Border(left=BORDER_SIDE, right=BORDER_SIDE, top=BORDER_SIDE, bottom=BORDER_SIDE)
CENTER      = Alignment(horizontal="center", vertical="center")
LEFT        = Alignment(horizontal="left", vertical="center")


def _apply_header(ws, columns):
    ws.row_dimensions[1].height = 30
    for ci, col in enumerate(columns, 1):
        cell = ws.cell(row=1, column=ci, value=col)
        cell.fill      = HEADER_FILL
        cell.font      = HEADER_FONT
        cell.border    = CELL_BORDER
        cell.alignment = CENTER
        ws.column_dimensions[cell.column_letter].width = 22


def make_registry_template_bytes() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Creator Registry"
    _apply_header(ws, REGISTRY_COLUMNS)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def make_unmatched_template_bytes(unmatched_ids: list) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Unmatched Creators"
    _apply_header(ws, UNMATCHED_COLUMNS)
    for ri, tid in enumerate(unmatched_ids, 2):
        ws.row_dimensions[ri].height = 20
        for ci, col in enumerate(UNMATCHED_COLUMNS, 1):
            val  = tid if col == "tiktok_id" else ""
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.font      = CELL_FONT
            cell.border    = CELL_BORDER
            cell.alignment = LEFT
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── helpers ───────────────────────────────────────────────────────────────────
def _parse_level(raw) -> int | None:
    if pd.isna(raw) or str(raw).strip() == "":
        return None
    m = re.search(r"\d+", str(raw))
    return int(m.group()) if m else None


def _parse_followers(raw) -> int | None:
    if pd.isna(raw) or str(raw).strip() == "":
        return None
    try:
        return int(float(str(raw).strip()))
    except ValueError:
        return None


def _val(row, col):
    v = row.get(col, "")
    return v if (pd.notna(v) and str(v).strip() != "") else None


# ── ONBOARDING IMPORTER ───────────────────────────────────────────────────────
def run_onboarding_importer():
    st.markdown("### 📅 Onboarding Date Importer")
    st.caption(
        "Upload an XLSX with **'Unique ID'**, **'Collaboration start time'**, "
        "**'Sales level'**, and **'Followers'** columns. "
        "Matched creators will have their date, level, followers, and binding status updated."
    )

    uploaded = st.file_uploader("Upload XLSX", type=["xlsx"], key="onboarding_upload")
    if not uploaded:
        return

    try:
        df = pd.read_excel(uploaded, dtype=str)
    except Exception as e:
        st.error(f"Failed to read file: {e}")
        return

    required = {"Unique ID", "Collaboration start time", "Sales level", "Followers"}
    missing  = required - set(df.columns)
    if missing:
        st.error(f"Missing required columns: {missing}")
        return

    df["_uid_norm"] = df["Unique ID"].str.strip().str.lower()
    df["_date_raw"] = pd.to_datetime(df["Collaboration start time"], errors="coerce")

    st.markdown(f"**Rows in file:** {len(df)}")

    if st.button("▶ Run Import", key="onboarding_run"):

        updated, skipped_bad_date = 0, 0

        try:
            conn = get_connection()
            with conn.cursor() as cur:
                # ── 1. Fetch existing tiktok_ids ─────────────────────────────
                cur.execute("SELECT tiktok_id FROM public.creator_registry")
                db_ids = {r["tiktok_id"].strip().lower() for r in cur.fetchall()}

                matched_rows   = df[df["_uid_norm"].isin(db_ids)]
                unmatched_rows = df[~df["_uid_norm"].isin(db_ids)]

                # ── 2. Update matched rows ────────────────────────────────────
                for _, row in matched_rows.iterrows():
                    level_val     = _parse_level(row.get("Sales level"))
                    followers_val = _parse_followers(row.get("Followers"))

                    if pd.isna(row["_date_raw"]):
                        skipped_bad_date += 1
                        cur.execute(
                            """
                            UPDATE public.creator_registry
                               SET binding_status = 'Bound',
                                   level          = %s,
                                   followers      = COALESCE(%s, followers)
                             WHERE LOWER(TRIM(tiktok_id)) = %s
                            """,
                            (level_val, followers_val, row["_uid_norm"]),
                        )
                    else:
                        date_val    = row["_date_raw"].date()
                        month_label = row["_date_raw"].strftime("%B %Y")
                        cur.execute(
                            """
                            UPDATE public.creator_registry
                               SET onboarding_date = %s,
                                   month_label     = %s,
                                   level           = %s,
                                   binding_status  = 'Bound',
                                   followers       = COALESCE(%s, followers)
                             WHERE LOWER(TRIM(tiktok_id)) = %s
                            """,
                            (date_val, month_label, level_val, followers_val, row["_uid_norm"]),
                        )
                        updated += 1
            conn.commit()
            conn.close()
        except Exception as e:
            st.error(f"Import failed: {e}")
            return

        # ── 3. Results ────────────────────────────────────────────────────────
        col1, col2, col3 = st.columns(3)
        col1.metric("✅ Updated",   updated)
        col2.metric("⚠️ Unmatched", len(unmatched_rows))
        col3.metric("🗓 Bad dates",  skipped_bad_date)

        if updated:
            st.success(f"{updated} creator(s) updated successfully.")

        if len(unmatched_rows):
            st.warning(
                f"{len(unmatched_rows)} ID(s) not found in DB. "
                "Download the template below, fill in the missing details, "
                "then upload it in **Import Unmatched** below."
            )
            unmatched_ids = unmatched_rows["Unique ID"].str.strip().tolist()
            xlsx_bytes    = make_unmatched_template_bytes(unmatched_ids)
            st.download_button(
                label="⬇ Download Unmatched Template",
                data=xlsx_bytes,
                file_name=f"unmatched_creators_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="dl_unmatched",
            )
            with st.expander("Preview unmatched IDs"):
                st.dataframe(
                    unmatched_rows[["Unique ID", "Collaboration start time", "Sales level"]],
                    use_container_width=True,
                )

    # ── Import Unmatched ──────────────────────────────────────────────────────
    st.divider()
    st.markdown("#### ➕ Import Unmatched Creators")
    st.caption(
        "Upload the filled unmatched template. "
        "The onboarding CSV above is used as a lookup for `followers`, `onboarding_date`, `level`, and `binding_status`."
    )

    uploaded_unmatched = st.file_uploader(
        "Upload filled unmatched template", type=["xlsx"], key="unmatched_import_upload"
    )
    if not uploaded_unmatched:
        return

    try:
        um_df = pd.read_excel(uploaded_unmatched, dtype=str)
    except Exception as e:
        st.error(f"Failed to read file: {e}")
        return

    um_missing = set(UNMATCHED_COLUMNS) - set(um_df.columns)
    if um_missing:
        st.error(f"Missing columns: {um_missing}. Please use the downloaded unmatched template.")
        return

    um_df["tiktok_id"] = um_df["tiktok_id"].str.strip()
    um_df = um_df[um_df["tiktok_id"].notna() & (um_df["tiktok_id"] != "")].copy()
    um_df["_tid_norm"] = um_df["tiktok_id"].str.lower()

    # Build lookup from onboarding CSV: tiktok_id_norm → onboarding row
    ob_lookup = df.set_index("_uid_norm")

    st.markdown(f"**Rows ready to insert:** {len(um_df)}")
    with st.expander("Preview"):
        st.dataframe(um_df[UNMATCHED_COLUMNS].head(20), use_container_width=True)

    if st.button("▶ Insert Unmatched into DB", key="unmatched_import_run"):

        try:
            conn = get_connection()
            with conn.cursor() as cur:
                cur.execute("SELECT tiktok_id FROM public.creator_registry")
                existing = {r["tiktok_id"].strip().lower() for r in cur.fetchall()}
            conn.close()
        except Exception as e:
            st.error(f"DB connection failed: {e}")
            return

        inserted, skipped_dup, skipped_no_ob = 0, 0, 0
        errors = []

        def _v(val):
            return val if (pd.notna(val) and str(val).strip() != "") else None

        try:
            conn = get_connection()
            with conn.cursor() as cur:
                for _, row in um_df.iterrows():
                    tid_norm = row["_tid_norm"]

                    if tid_norm in existing:
                        skipped_dup += 1
                        continue

                    # Look up onboarding data
                    if tid_norm in ob_lookup.index:
                        ob_row      = ob_lookup.loc[tid_norm]
                        date_raw    = ob_row["_date_raw"] if "_date_raw" in ob_row.index else pd.NaT
                        date_val    = date_raw.date() if pd.notna(date_raw) else None
                        month_label = date_raw.strftime("%B %Y") if pd.notna(date_raw) else None
                        level_val   = _parse_level(ob_row.get("Sales level"))
                        followers_v = _parse_followers(ob_row.get("Followers"))
                        b_status    = "Bound"
                    else:
                        skipped_no_ob += 1
                        date_val = month_label = level_val = followers_v = None
                        b_status = "Unbound"

                    tiktok_link = f"https://www.tiktok.com/@{row['tiktok_id']}"

                    try:
                        cur.execute(
                            """
                            INSERT INTO public.creator_registry
                                (tiktok_id, followers, full_name, domicile, uid,
                                 phone_number, tiktok_link, binding_status,
                                 onboarding_date, month_label, level, agency_id, notes)
                            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                            """,
                            (
                                row["tiktok_id"],
                                followers_v,
                                _v(row.get("full_name")),
                                _v(row.get("domicile")),
                                _v(row.get("uid")),
                                _v(row.get("phone_number")),
                                tiktok_link,
                                b_status,
                                date_val,
                                month_label,
                                level_val,
                                1,
                                None,
                            ),
                        )
                        inserted += 1
                    except Exception as row_err:
                        errors.append(f"Row {row['tiktok_id']}: {row_err}")

            conn.commit()
            conn.close()
        except Exception as e:
            st.error(f"Insert failed: {e}")
            return

        col1, col2, col3, col4 = st.columns(4)
        col1.metric("✅ Inserted",        inserted)
        col2.metric("⏭ Duplicates",      skipped_dup)
        col3.metric("⚠️ No onboard data", skipped_no_ob)
        col4.metric("❌ Errors",          len(errors))

        if inserted:
            st.success(f"{inserted} creator(s) inserted successfully.")
        if skipped_no_ob:
            st.info(f"{skipped_no_ob} row(s) inserted without onboarding data — not found in the onboarding CSV.")
        if errors:
            with st.expander("Error details"):
                for err in errors:
                    st.error(err)


# ── SHARED BULK IMPORTER ──────────────────────────────────────────────────────
def _run_bulk_importer(agency_id: int, key_prefix: str, success_suffix: str = ""):
    st.download_button(
        label="⬇ Download Blank Template",
        data=make_registry_template_bytes(),
        file_name=f"{key_prefix}_registry_template.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        key=f"dl_{key_prefix}_template",
    )

    uploaded = st.file_uploader("Upload filled template", type=["xlsx"], key=f"{key_prefix}_upload")
    if not uploaded:
        return

    try:
        df = pd.read_excel(uploaded, dtype=str)
    except Exception as e:
        st.error(f"Failed to read file: {e}")
        return

    missing = set(REGISTRY_COLUMNS) - set(df.columns)
    if missing:
        st.error(f"Template columns missing: {missing}. Please use the provided template.")
        return

    df = df[REGISTRY_COLUMNS].copy()
    df["tiktok_id"] = df["tiktok_id"].str.strip()
    df = df[df["tiktok_id"].notna() & (df["tiktok_id"] != "")]
    df["_date"] = pd.to_datetime(df["onboarding_date"], errors="coerce")

    st.markdown(f"**Rows ready to insert:** {len(df)}")

    bad_dates = df[df["_date"].isna() & df["onboarding_date"].notna() & (df["onboarding_date"] != "")]
    if len(bad_dates):
        st.warning(f"{len(bad_dates)} row(s) have unparseable dates — will be inserted with NULL onboarding_date.")

    with st.expander("Preview data"):
        st.dataframe(df[REGISTRY_COLUMNS].head(20), use_container_width=True)

    if st.button("▶ Insert into DB", key=f"{key_prefix}_run"):

        inserted, skipped_dup = 0, 0
        errors = []

        try:
            conn = get_connection()
            with conn.cursor() as cur:
                # ── 1. Fetch existing ids ─────────────────────────────────────
                cur.execute("SELECT tiktok_id FROM public.creator_registry")
                existing = {row["tiktok_id"].strip().lower() for row in cur.fetchall()}

                # ── 2. Insert ─────────────────────────────────────────────────
                for _, row in df.iterrows():
                    if row["tiktok_id"].lower() in existing:
                        skipped_dup += 1
                        continue

                    date_val    = row["_date"].date() if pd.notna(row["_date"]) else None
                    month_label = row["_date"].strftime("%B %Y") if pd.notna(row["_date"]) else None

                    try:
                        cur.execute(
                            """
                            INSERT INTO public.creator_registry
                                (tiktok_id, followers, full_name, domicile, uid,
                                 phone_number, tiktok_link, binding_status,
                                 onboarding_date, month_label, agency_id, notes)
                            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                            """,
                            (
                                row["tiktok_id"],
                                _val(row, "followers"),
                                _val(row, "full_name"),
                                _val(row, "domicile"),
                                _val(row, "uid"),
                                _val(row, "phone_number"),
                                _val(row, "tiktok_link"),
                                _val(row, "binding_status"),
                                date_val,
                                month_label,
                                agency_id,
                                None,
                            ),
                        )
                        inserted += 1
                    except Exception as row_err:
                        errors.append(f"Row {row['tiktok_id']}: {row_err}")

            conn.commit()
            conn.close()
        except Exception as e:
            st.error(f"Insert failed: {e}")
            return

        # ── 3. Results ────────────────────────────────────────────────────────
        col1, col2, col3 = st.columns(3)
        col1.metric("✅ Inserted",   inserted)
        col2.metric("⏭ Duplicates", skipped_dup)
        col3.metric("❌ Errors",     len(errors))

        if inserted:
            st.success(f"{inserted} creator(s) inserted{success_suffix}.")
        if skipped_dup:
            st.info(f"{skipped_dup} row(s) skipped — tiktok_id already exists in DB.")
        if errors:
            with st.expander("Error details"):
                for err in errors:
                    st.error(err)


# ── CREATOR REGISTRY IMPORTER ─────────────────────────────────────────────────
def run_registry_importer():
    st.markdown("### 👤 Creator Registry Importer")
    st.caption(
        "Use this tab to insert **new** creators into the DB. "
        "Download the template, fill in the details, then upload. "
        "Optionally upload the onboarding CSV to auto-fill UID from `author_id`."
    )
    _run_bulk_importer(agency_id=1, key_prefix="registry")


# ── VENDOR IMPORT ────────────────────────────────────────────────────────────
def run_vendor_importer():
    st.markdown("### 🏢 Vendor Import")
    st.caption(
        "Import creators from a vendor CSV. "
        "Select the vendor/agency first — all rows in the file will be assigned to that agency."
    )

    try:
        conn = get_connection()
        with conn.cursor() as cur:
            cur.execute("SELECT id, agency_name FROM public.agency_map ORDER BY id")
            agencies = cur.fetchall()
        conn.close()
    except Exception as e:
        st.error(f"Failed to load agency list: {e}")
        return

    if not agencies:
        st.error("No agencies found. Please add agencies in the Settings page first.")
        return

    agency_options = {f"{a['id']} — {a['agency_name']}": a["id"] for a in agencies}
    choice    = st.selectbox("Select Vendor / Agency", list(agency_options.keys()), key="vendor_agency")
    agency_id = agency_options[choice]

    _run_bulk_importer(agency_id=agency_id, key_prefix="vendor", success_suffix=f" under agency '{choice}'")


# ── SANITY CHECK ──────────────────────────────────────────────────────────────
def run_sanity_check():
    st.markdown("### 🔍 Sanity Check")
    st.caption(
        "Upload the onboarding CSV to find creators in the DB whose `tiktok_id` "
        "does **not** appear as `Unique ID` in the file."
    )

    uploaded = st.file_uploader("Upload onboarding XLSX", type=["xlsx"], key="sanity_upload")
    if not uploaded:
        return

    try:
        df = pd.read_excel(uploaded, dtype=str)
    except Exception as e:
        st.error(f"Failed to read file: {e}")
        return

    if "Unique ID" not in df.columns:
        st.error("Missing 'Unique ID' column.")
        return

    csv_ids = set(df["Unique ID"].str.strip().str.lower().dropna())

    try:
        conn = get_connection()
        with conn.cursor() as cur:
            cur.execute(
                """
                SELECT cr.tiktok_id, cr.full_name, am.agency_name
                  FROM public.creator_registry cr
                  LEFT JOIN public.agency_map am ON cr.agency_id = am.id
                 ORDER BY cr.tiktok_id
                """
            )
            db_rows = cur.fetchall()
        conn.close()
    except Exception as e:
        st.error(f"DB connection failed: {e}")
        return

    missing = [
        {"tiktok_id": r["tiktok_id"], "full_name": r["full_name"], "agency": r["agency_name"]}
        for r in db_rows
        if r["tiktok_id"].strip().lower() not in csv_ids
    ]

    col1, col2 = st.columns(2)
    col1.metric("Creators in DB", len(db_rows))
    col2.metric("Not in CSV", len(missing))

    if missing:
        st.warning(f"{len(missing)} creator(s) in DB are missing from the onboarding CSV.")
        st.dataframe(missing, use_container_width=True)
    else:
        st.success("All DB creators are present in the CSV.")


# ── PAGE ──────────────────────────────────────────────────────────────────────
def render():
    st.title("🎬 Creator Importer")

    tab_onboarding, tab_registry, tab_vendor, tab_sanity = st.tabs([
        "📅 Onboarding Date", "👤 Creator Registry", "🏢 Vendor Import", "🔍 Sanity Check"
    ])

    with tab_onboarding:
        run_onboarding_importer()

    with tab_registry:
        run_registry_importer()

    with tab_vendor:
        run_vendor_importer()

    with tab_sanity:
        run_sanity_check()


if __name__ == "__main__":
    render()
